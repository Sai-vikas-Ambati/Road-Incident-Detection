import requests
import json
import openpyxl
import time
import os

# Configurable settings
EXCEL_FILE = "traffic_data.xlsx"   # General name instead of Sept6.xlsx
SHEET_NAME = "TrafficLog"
API_KEY = os.getenv("GOOGLE_MAPS_API_KEY")  # Read API key from environment variable
ORIGIN = "17.443736358962692,78.47755695419899"
DESTINATION = "17.444236001504468,78.47097376785939"

# Load or create Excel workbook
try:
    workbook = openpyxl.load_workbook(EXCEL_FILE)
    worksheet = workbook.active
except FileNotFoundError:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    worksheet.cell(row=1, column=1, value="normal_traffic")
    worksheet.cell(row=1, column=2, value="traffic_delay")
    worksheet.cell(row=1, column=3, value="TIME")

def main(i):
    worksheet = workbook.active
    url = (
        f"https://maps.googleapis.com/maps/api/directions/json?"
        f"origin={ORIGIN}&destination={DESTINATION}"
        f"&key={API_KEY}&departure_time=now&traffic_model=best_guess"
    )
    response = requests.get(url)
    data = json.loads(response.text)

    try:
        normal_time = data["routes"][0]["legs"][0]["duration"]["value"]
        traffic_time = data["routes"][0]["legs"][0]["duration_in_traffic"]["value"]

        worksheet.cell(row=i, column=1, value=normal_time)
        worksheet.cell(row=i, column=2, value=traffic_time)
        worksheet.cell(row=i, column=3, value=time.strftime("%H:%M:%S"))
        workbook.save(EXCEL_FILE)
    except (KeyError, IndexError):
        print("⚠️ Could not extract data from API response")

end_time = time.time() + 60 * 60 * 24  # Run for 24 hours
i = (worksheet.max_row) + 1

while time.time() <= end_time:
    main(i)
    i += 1
    time.sleep(2 * 60)  # Run every 2 minutes
